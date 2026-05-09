# -*- coding: utf-8 -*-
"""
Created on Mon Feb  5 09:45:48 2024

@author: kwater
"""

import psutil
import os
import sys
import logging
import pandas as pd
from glob import glob
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.chrome.options import Options
from concurrent.futures import ThreadPoolExecutor
from functools import partial
from queue import Queue
from concurrent.futures import ProcessPoolExecutor

DRIVER_PATH = "./chromedriver-win64/chromedriver.exe"
ALL_ELEMENTS = []
ALL_SN = []

def shutdown() :
    os.system("shutdown -s -t 0")
    
def reboot():
    os.system("shutdown -r -t 0")

def memoryCheck():
    return psutil.virtual_memory().percent

def addSN(SN):
    ALL_SN.extend(SN)
    
def getSN():
    return ALL_SN

def addElements(elements):
    ALL_ELEMENTS.extend(elements)
    
def getElements():
    return ALL_ELEMENTS

def setChromeDriver():
    # set chrome driver path
    chrome_driver_path = DRIVER_PATH

    # set chrome option
    chrome_options = Options()
    chrome_options.add_argument("--headless")  # web browser invisible

    # Create Chrome service
    chrome_service = ChromeService(executable_path=chrome_driver_path)

    # create chrome driver
    driver = webdriver.Chrome(service=chrome_service, options=chrome_options)
    
    return driver

def castSN(attributes):
    return [attr.replace("javaScript:fn_detail(this,'", "").replace("','taol');", "").replace("','tqapttn');", "") for attr in attributes]

def castDetailURL(SN):
    detail_urls = []

    for sn in castSN(SN):
        if len(sn) > 10:
            detail_url = f'https://www.epeople.go.kr/nep/pttn/gnrlPttn/pttnSmlrCaseDetail.npaid?epUnionSn={sn}&dutySctnNm=taol'
        else:
            detail_url = f'https://www.epeople.go.kr/nep/pttn/gnrlPttn/pttnSmlrCaseDetail.npaid?epUnionSn={sn}&dutySctnNm=tqapttn'

        detail_urls.append(detail_url)

    return detail_urls

def crawling_SN(base_url, crawling_page):
    attributes = []
    elements = []

    driver = setChromeDriver()
    
    # open web browser
    url = base_url + f"&pageIndex={crawling_page}"
    driver.get(url)

    # wait 5 seconds for page loading
    driver.implicitly_wait(5)

    # read page source
    page_source = driver.page_source
    
    # abstract contents in page source(tag:a / class:tit) and print desc
    for element in driver.find_elements(By.XPATH, "//a[@class='tit']"):
        elements.append(element.text)
        print(element.text)
    
    soup = BeautifulSoup(page_source, "html.parser")     # seperate by html code type

    for i, a_tag in enumerate(soup.find_all("a", {'class':"tit"})):
        attributes.append(a_tag["onclick"])
        
    # terminate driver
    driver.quit()

    return attributes, elements

def process_url(url, result_queue):
    if memoryCheck() >= 92.0:
        print("No enough memory")
        sys.exit()
    
    top_contents = []
    info_contents = []
    
    # config logging
    logging.basicConfig(filename='scraping_errors.log', level=logging.ERROR)
    
    # js code for excuting chrome browser garbage collector
    js_code = """
    if (window.gc) {
            window.gc();
    } else {
        console.warn("Garbage collection not available. Check browser compatibility.");
    }
    """
    
    try:
        # create chrome driver
        driver = setChromeDriver()
        
        driver.get(url)

        # wait 5 seconds for page loading
        driver.implicitly_wait(5)

        # read page source
        soup = BeautifulSoup(driver.page_source, "html.parser")  # separate by html code type
        
        for strong_tag in soup.find_all("strong"):
            strong_tag.decompose()

        for i, a_tag in enumerate(set(soup.find_all("div", {'class': "samC_top"}))):
            top_contents.append(a_tag)
        
        top_contents = [content for content in top_contents if content]

        for i, a_tag in enumerate(set(soup.find_all("div", {'class': "samC_c"}))):
            info_contents.append(a_tag)
            
        driver.execute_script(js_code)
            
        # terminate driver
        driver.quit()

        print('...')
        
        return top_contents, info_contents

    except Exception as e:
        logging.error(f"An unexpected error occurred: {e}")
        return None, None

def crawling_txt(SN):
    detail_urls = []
    top_contents = []
    info_contents = []

    detail_urls = castDetailURL(SN)

    print(detail_urls)
    
    # Create a Queue to collect results
    result_queue = Queue()
    
    # Partially apply process_url with result_queue
    partial_process_url = partial(process_url, result_queue=result_queue)
    
    with ThreadPoolExecutor() as executor:
        # Use executor.map to process URLs and collect results
        results = list(executor.map(partial_process_url, detail_urls))
        # Unpack the results into separate lists
        top_contents, info_contents = zip(*results)
        #top_contents, info_contents = list(executor.map(partial_process_url, detail_urls))
    
    return top_contents, info_contents

def save(elements, top_contents, info_contents, crawling_page=-1):
    wb = Workbook()
    sheet = wb.active
    file_name = 'complaints' + str(crawling_page) + '.xlsx'

    sheet.column_dimensions["A"].width = 6
    sheet.column_dimensions["B"].width = 110
    sheet.column_dimensions["C"].width = 130
    
    sheet.cell(row=1, column=1).value = 'Index'
    sheet.cell(row=1, column=2).value = 'Title'
    sheet.cell(row=1, column=3).value = 'Department / Related_Law'
    sheet.cell(row=1, column=4).value = 'Answer'

    for idx in range(0, len(elements)):
        # cell styling
        alignment = Alignment(horizontal="center", vertical="center")
        font = Font(name="Arial", color=Color(rgb="000000"))
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
        
        cell_range = sheet[f"A{1}:D{len(elements) + 1}"]   # cell range for styling
        
        for row in cell_range:
            for cell in row:
                cell.alignment = alignment
                cell.font = font
                cell.border = border
        
        # input data
        sheet.cell(row=idx + 2, column=1).value = idx + 1
        sheet.cell(row=idx + 2, column=2).value = elements[idx]
        sheet.cell(row=idx + 2, column=3).value = str(info_contents[idx])
        sheet.cell(row=idx + 2, column=4).value = str(top_contents[idx])
        #sheet.cell(row=idx + 2, column=3).value = info_contents[idx].get_text()
        #sheet.cell(row=idx + 2, column=4).value = top_contents[idx].get_text()

    wb.save(filename=file_name)

def crawling(base_url, maxPageCount, recordCountPerPage):
    startPage = 1
    
    for crawling_page in range(startPage, 1 + maxPageCount):
        
        # attr crawling
        attributes, elements = crawling_SN(base_url, crawling_page)
        print('=======================================')
        
        # add elements to ALL ELEMENTS
        addElements(elements)
        
        # abstract serial numbers in attributes
        SN = castSN(attributes)
        print('=======================================')
        
        # add SN to ALL_SN
        addSN(SN)
        """
        # cont crawling
        top_contents, info_contents = crawling_txt(SN)
        print('=====================================================')
        
        # contents save by excel file
        print(top_contents, info_contents)
        #save(elements, top_contents, info_contents, crawling_page)
        print('=====================================================')
        """
        
def mergeSave():
    # get cwd and replace backspace to slash
    current_dir = os.getcwd().replace('\\', '/')
    # set base_url to cwd
    base_dir = f'{current_dir}/'
    files = glob(os.path.join(base_dir, '*.xlsx'))
    mergedDataFrame = pd.DataFrame()
    
    for file in files:
        temp = pd.read_excel(file)
        mergedDataFrame = pd.concat([mergedDataFrame, temp])

    mergedDataFrame.to_excel('complaints.xlsx', index=False)

def getMaxPage(url):
    maxPageCount = 1

    # create chrome driver
    driver = setChromeDriver()
    
    # open web browser
    driver.get(url)

    # wait 5 seconds for page loading
    driver.implicitly_wait(5)
    
    # read page source
    page_source = driver.page_source
    
    soup = BeautifulSoup(page_source, "html.parser")     # seperate by html code type

    span_tag = soup.find("span", {'class':"paging_count"})
    
    maxPageCount = int(str(span_tag).replace('<span class="paging_count">1/', '').replace('</span>', ''))
    
    # terminate driver
    driver.quit()
    
    return maxPageCount



def processDelDfCrawling(delt, base_url, result_queue):
    if memoryCheck() >= 92.0:
        print("No enough memory")
        sys.exit()
    
    attributes = []
    elements = []
    
    # config logging
    logging.basicConfig(filename='scraping_errors.log', level=logging.ERROR)
    
    # js code for excuting chrome browser garbage collector
    js_code = """
    if (window.gc) {
            window.gc();
    } else {
        console.warn("Garbage collection not available. Check browser compatibility.");
    }
    """
    
    driver = setChromeDriver()
    
    # open web browser
    url = base_url
    driver.get(url)

    # wait 2 seconds for page loading
    driver.implicitly_wait(2)
    
    # click and input search_word
    driver.find_element(By.ID, 'searchWord').click()
    driver.find_element(By.ID, 'searchWord').send_keys(delt)
    #driver.find_element_by_id('searchWord').click()
    #driver.find_element_by_id('searchWord').send_keys(delete_df['Title'])
    
    # execute search by clicking search button
    driver.find_element(By.XPATH, "//button[@class='btn black']").click()
    #driver.find_element_by_class_name('btn black').click()
    
    # read page source
    page_source = driver.page_source
    
    # abstract contents in page source(tag:a / class:tit) and print desc
    for element in driver.find_elements(By.XPATH, "//a[@class='tit']"):
        elements.append(element.text)
        print(element.text)
    
    soup = BeautifulSoup(page_source, "html.parser")     # seperate by html code type

    for i, a_tag in enumerate(soup.find_all("a", {'class':"tit"})):
        attributes.append(a_tag["onclick"])
        
    driver.execute_script(js_code)
        
    # terminate driver
    driver.quit()
    
    return attributes, elements
"""
def delDfCrawling(delt, base_url):
    # Create a Queue to collect results
    result_queue = Queue()
    
    # Partially apply process_url with result_queue
    partial_del_process_url = partial(processDelDfCrawling, result_queue=result_queue)
    
    with ThreadPoolExecutor() as executor:
        # Use executor.map to process URLs and collect results
        results = list(executor.map(partial_del_process_url, delt, base_url))
        # Unpack the results into separate lists
        attributes, elements = zip(*results)
        
    return attributes, elements
"""

def delDfCrawling(delt, base_url):
    attributes = []
    elements = []
    
    driver = setChromeDriver()
    
    # open web browser
    url = base_url
    driver.get(url)

    # wait 2 seconds for page loading
    driver.implicitly_wait(2)
    
    # click and input search_word
    driver.find_element(By.ID, 'searchWord').click()
    driver.find_element(By.ID, 'searchWord').send_keys(delt)
    #driver.find_element_by_id('searchWord').click()
    #driver.find_element_by_id('searchWord').send_keys(delete_df['Title'])
    
    # execute search by clicking search button
    driver.find_element(By.XPATH, "//button[@class='btn black']").click()
    #driver.find_element_by_class_name('btn black').click()
    
    # read page source
    page_source = driver.page_source
    
    # abstract contents in page source(tag:a / class:tit) and print desc
    for element in driver.find_elements(By.XPATH, "//a[@class='tit']"):
        elements.append(element.text)
        print(element.text)
    
    soup = BeautifulSoup(page_source, "html.parser")     # seperate by html code type

    for i, a_tag in enumerate(soup.find_all("a", {'class':"tit"})):
        attributes.append(a_tag["onclick"])
        
    # terminate driver
    driver.quit()
    
    print('...')
    
    return attributes, elements
"""
def process_delDfCrawling(title, base_url):
    return delDfCrawling(title, base_url)
"""

def naCheck(base_url):
    # get cwd and replace backspace to slash
    current_dir = os.getcwd().replace('\\', '/')
    # set base_url to cwd
    base_dir = f'{current_dir}/'
    file = 'complaints.xlsx'
    
    df = pd.read_excel(os.path.join(base_dir, file))
    
    delete_df = df[df['Department / Related_Law'] == '[]']
    modified_df = df[df['Department / Related_Law'] != '[]']
    
    print(modified_df)
    print(delete_df)
    
    attributes, elements = [], []
    
    """
    # Using ProcessPoolExecutor to parallelize delDfCrawling function
    with ProcessPoolExecutor() as executor:
        results = list(executor.map(process_delDfCrawling, delete_df['Title'], [base_url]*len(delete_df)))

    for result in results:
        attributes.extend(result[0])
        elements.extend(result[1])
    """
    
    """
    for delt in delete_df['Title']:
        attr, e = delDfCrawling(delt, base_url)
        attributes.extend(attr)
        elements.extend(e)
    
    SN = castSN(attributes)
    
    top_contents, info_contents = crawling_txt(SN)
    print('=====================================================')
    
    # contents save by excel file
    print(top_contents, info_contents)
    save(elements, top_contents, info_contents)
    """
    
def dropBlank():
    # get cwd and replace backspace to slash
    current_dir = os.getcwd().replace('\\', '/')
    # set base_url to cwd
    base_dir = f'{current_dir}/'
    file = 'complaints-1.xlsx'
    
    df = pd.read_excel(os.path.join(base_dir, file))
    
    delete_df = df[df['Department / Related_Law'] != '[]']
    
    delete_df.drop('Index', axis=1)
    
    delete_df.to_excel('complaints_f.xlsx', index=False)
    
def dropDuplicates():
    # get cwd and replace backspace to slash
    current_dir = os.getcwd().replace('\\', '/')
    # set base_url to cwd
    base_dir = f'{current_dir}/'
    file = 'complaints.xlsx'
    
    df = pd.read_excel(os.path.join(base_dir, file))
    
    non_duplicated_df = df.drop_duplicates(['Department / Related_Law', 'Answer'])
    
    non_duplicated_df.to_excel('complaints_d.xlsx', index=False)
    
def main(base_url, maxPageCount, recordCountPerPage):
    #crawling(base_url, maxPageCount, recordCountPerPage)
    #mergeSave()
    #dropBlank()
    naCheck(base_url)
    #dropBlank()
    #dropDuplicates()
    
    
if __name__ == "__main__":
    recordCountPerPage = 26273
    base_url = f'https://www.epeople.go.kr/nep/pttn/gnrlPttn/pttnSmlrCaseList.npaid?recordCountPerPage={recordCountPerPage}'
    maxPageCount = getMaxPage(base_url)
    main(base_url, maxPageCount, recordCountPerPage)
    
